"""File handling utilities for the concierge application."""

# File Upload Configuration
ALLOWED_EXTENSIONS = {'txt', 'pdf', 'docx', 'xlsx'}

def allowed_file(filename):
    """Check if a filename has an allowed extension.
    
    Args:
        filename: The name of the file to check
        
    Returns:
        bool: True if the file extension is allowed, False otherwise
    """
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS 

# --- Text Extraction Helpers ---
def extract_text_from_pdf(file_path):
    """Extract text content from a PDF file.
    
    Args:
        file_path: Path to the PDF file
        
    Returns:
        str: Extracted text content
    """
    text = ""
    try:
        import pypdf
        reader = pypdf.PdfReader(file_path)
        for page in reader.pages:
            text += page.extract_text() or ""
    except Exception as e:
        print(f"Error extracting text from PDF {file_path}: {e}")
    return text

def extract_text_from_docx(file_path):
    """Extract text content from a DOCX file.
    
    Args:
        file_path: Path to the DOCX file
        
    Returns:
        str: Extracted text content
    """
    text = ""
    try:
        import docx
        doc = docx.Document(file_path)
        for para in doc.paragraphs:
            text += para.text + "\n"
    except Exception as e:
        print(f"Error extracting text from DOCX {file_path}: {e}")
    return text

def extract_text_from_xlsx(file_path):
    """Extract text content from an XLSX file.
    
    Args:
        file_path: Path to the XLSX file
        
    Returns:
        str: Extracted text content
    """
    text = ""
    try:
        import openpyxl
        # Load the workbook
        wb = openpyxl.load_workbook(file_path, data_only=True)
        # Process each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Process each row
            for row in ws.iter_rows():
                # Process each cell in the row
                for cell in row:
                    if cell.value is not None:
                        text += str(cell.value) + "\n"
            text += "\n\n"  # Separate sheets
    except Exception as e:
        print(f"Error extracting text from XLSX {file_path}: {e}")
    return text

def extract_text_from_txt(file_path):
    """Extract text content from a plain text file.
    
    Args:
        file_path: Path to the text file
        
    Returns:
        str: Extracted text content
    """
    text = ""
    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            text = f.read()
    except Exception as e:
        print(f"Error reading text file {file_path}: {e}")
    return text 